from openpyxl import load_workbook

def update_file(today):
    workbook1 = load_workbook('Excel_Copy.xlsx')

    workbook2 = load_workbook('Excel_testing.xlsx')

    sheet_to_replace = workbook1[today]
    replacement_sheet = workbook2[today]
    sheet_to_replace.delete_rows(1, sheet_to_replace.max_row)
    for row in replacement_sheet.iter_rows(values_only=True):
        sheet_to_replace.append(row)

    workbook1.save('Excel_copy.xlsx')
